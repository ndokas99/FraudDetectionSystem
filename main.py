from openpyxl import load_workbook

workbook = load_workbook("input/input.xlsx")
worksheet = workbook.active
data = {}

for row in worksheet.iter_rows(values_only=True):
    data[row[0]] = row[1]

acr = data["Opening Accounts Receivable"]
clr = data["Closing Accounts Receivable"]
crs = data["Credit Sales"]
tos = data["Total Sales"]
cogs = data["Cost of Goods Sold"]
cua = data["Current Assets"]
cas = data["Cash"]
fia = data["Fixed Assets"]
toa = cua + fia
pys = data["Previous Years Sales"]
dep = data["Depreciation Expense"]
amr = data["Amortization Expense"]
see = data["Selling Expenses"]
gex = data["General Expenses"]
ade = data["Administration expenses"]
tol = data["Total Liabilities"]
nei = data["Net Income"]
cfo = data["Cashflow from operations"]

DSRI = (((acr+clr)/2) / (crs / 365))
GMI = ((tos - cogs) / tos)
AQI = 1 - (cua - cas - clr) / toa
SGI = (tos - pys) / pys
DEPI = (dep + amr) / toa
SGAI = 1 if (see + gex + ade - dep - amr) / toa > 1 else 0
LVGI = 1 if tol / toa > 1 else 0
TATA = 1 if (nei - cfo) / toa > 0 else 0

Beneish_M_Score = -4.84 + 0.92 * DSRI + 0.528 * GMI + 0.404 * AQI + 0.892 * SGI \
                  + 0.115 * DEPI + 0.172 * SGAI + 4.679 * TATA - 0.327 * LVGI


with open("output/results.txt", "w") as file:
    if Beneish_M_Score > 6.81:
        file.write(f"""__Results of analysis__
        
Beneish M-Score reported a result of {Beneish_M_Score:.3} which is higher than 6.81
expected of the model thus there is a high chance of manipulation of financial statements.    
""")
    elif Beneish_M_Score < -4.84:
        file.write(f"""__Results of analysis__

Beneish M-Score reported a result of {Beneish_M_Score:.3} which is lower than -4.84
expected of the model thus there is a very low chance of manipulation of financial statements.    
        """)
    else:
        file.write(f"""__Results of analysis__

Beneish M-Score reported a result of {Beneish_M_Score:.3} which is between -4.84 and
6.81 expected of the model thus the company seems to be operating in a risky zone.    
        """)

